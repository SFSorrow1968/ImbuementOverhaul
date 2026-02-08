#!/usr/bin/env python
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
DOCS_DIR = ROOT / "_docs"
MENU_PATH = DOCS_DIR / "EnemyImbuePresets_MenuMock.xlsx"
MATRIX_PATH = DOCS_DIR / "EnemyImbuePresets_PresetMatrix.xlsx"


FACTIONS = [
    ("combat", "Combat Practice (Mixed Enemies)", "Combat"),
    ("outlaws", "Outlaws (T0)", "Outlaws"),
    ("wildfolk", "Wildfolk (T1)", "Wildfolk"),
    ("eraden", "Eraden Kingdom (T2)", "Eraden"),
    ("eye", "The Eye (T3)", "Eye"),
    ("rakta", "Rakta", "Rakta"),
    ("special", "Special / Rogue", "Special"),
    ("fallback", "Fallback (Any Enemy)", "Fallback"),
]

FACTION_PROFILE_PRESETS = [
    ("LoreFriendly", "Default"),
    ("FrontierPressure", "Core Factions"),
    ("WarfrontArcana", "Most Factions"),
    ("HighMagicConflict", "All Factions"),
    ("RandomizedEligibility", "Random"),
]

ENEMY_TYPE_PROFILE_PRESETS = [
    ("LoreFriendly", "Mage"),
    ("FrontierPressure", "Mage Bow"),
    ("WarfrontArcana", "Mage Melee"),
    ("HighMagicConflict", "Mage Bow Melee"),
    ("RandomizedEligibility", "Random"),
]

IMBUE_PRESETS = [
    ("LoreAccurate", "Default"),
    ("FactionIdentity", "Two-Slot"),
    ("ArcaneSurge", "Tri-Slot"),
    ("ElementalChaos", "Tri-Slot+"),
    ("Randomized", "Random"),
]

CHANCE_PRESETS = [
    ("LowIntensity", "Default"),
    ("BalancedBattles", "Increased"),
    ("AggressiveWaves", "High"),
    ("RelentlessThreat", "Very High"),
    ("OverflowNormalized", "Maximum"),
]

STRENGTH_PRESETS = [
    ("FaintCharge", "Default"),
    ("BattleReady", "Increased"),
    ("Empowered", "High"),
    ("Overcharged", "Very High"),
    ("Cataclysmic", "Maximum"),
]


FACTION_PROFILE_VALUES = {
    "LoreFriendly": [False, False, False, True, True, True, True, True],
    "FrontierPressure": [False, False, True, True, True, True, True, True],
    "WarfrontArcana": [False, True, True, True, True, True, True, True],
    "HighMagicConflict": [True, True, True, True, True, True, True, True],
}

ENEMY_TYPE_PROFILE_VALUES = {
    "LoreFriendly": [True, True, True, False, False],
    "FrontierPressure": [True, True, True, True, False],
    "WarfrontArcana": [True, True, True, False, True],
    "HighMagicConflict": [True, True, True, True, True],
}

IMBUE_VALUES = {
    "LoreAccurate": [
        ["None", "None", "None"],
        ["None", "None", "None"],
        ["None", "None", "None"],
        ["Fire", "None", "None"],
        ["Gravity", "None", "None"],
        ["Lightning", "None", "None"],
        ["Gravity", "None", "None"],
        ["Fire", "None", "None"],
    ],
    "FactionIdentity": [
        ["None", "None", "None"],
        ["None", "None", "None"],
        ["Lightning", "None", "None"],
        ["Fire", "Lightning", "None"],
        ["Gravity", "Lightning", "None"],
        ["Fire", "Lightning", "None"],
        ["Gravity", "Fire", "None"],
        ["Fire", "Lightning", "None"],
    ],
    "ArcaneSurge": [
        ["Fire", "None", "None"],
        ["Fire", "None", "None"],
        ["Lightning", "Fire", "None"],
        ["Fire", "Lightning", "Gravity"],
        ["Gravity", "Lightning", "Fire"],
        ["Lightning", "Fire", "Gravity"],
        ["Gravity", "Fire", "Lightning"],
        ["Fire", "Lightning", "Gravity"],
    ],
    "ElementalChaos": [
        ["Fire", "Lightning", "None"],
        ["Fire", "Lightning", "None"],
        ["Lightning", "Fire", "None"],
        ["Fire", "Lightning", "Gravity"],
        ["Gravity", "Lightning", "Fire"],
        ["Lightning", "Fire", "Gravity"],
        ["Gravity", "Fire", "Lightning"],
        ["Fire", "Lightning", "Gravity"],
    ],
}

CHANCE_VALUES = {
    "LowIntensity": [
        [0.0, 0.0, 0.0],
        [0.0, 0.0, 0.0],
        [6.0, 0.0, 0.0],
        [24.0, 0.0, 0.0],
        [28.0, 0.0, 0.0],
        [30.0, 0.0, 0.0],
        [16.0, 0.0, 0.0],
        [12.0, 0.0, 0.0],
    ],
    "BalancedBattles": [
        [2.0, 0.0, 0.0],
        [4.0, 0.0, 0.0],
        [10.0, 2.0, 0.0],
        [30.0, 14.0, 0.0],
        [34.0, 16.0, 0.0],
        [36.0, 18.0, 0.0],
        [20.0, 8.0, 0.0],
        [16.0, 6.0, 0.0],
    ],
    "AggressiveWaves": [
        [6.0, 2.0, 0.0],
        [8.0, 4.0, 0.0],
        [14.0, 6.0, 0.0],
        [34.0, 18.0, 8.0],
        [38.0, 20.0, 10.0],
        [40.0, 22.0, 10.0],
        [24.0, 12.0, 6.0],
        [20.0, 10.0, 4.0],
    ],
    "RelentlessThreat": [
        [10.0, 4.0, 0.0],
        [12.0, 6.0, 0.0],
        [18.0, 8.0, 2.0],
        [38.0, 22.0, 10.0],
        [42.0, 24.0, 12.0],
        [44.0, 26.0, 12.0],
        [28.0, 14.0, 8.0],
        [24.0, 12.0, 6.0],
    ],
    "OverflowNormalized": [
        [14.0, 8.0, 2.0],
        [18.0, 10.0, 4.0],
        [24.0, 14.0, 6.0],
        [46.0, 28.0, 14.0],
        [50.0, 30.0, 16.0],
        [52.0, 32.0, 16.0],
        [34.0, 18.0, 10.0],
        [30.0, 16.0, 8.0],
    ],
}

STRENGTH_VALUES = {
    "FaintCharge": [
        [0.0, 0.0, 0.0],
        [0.0, 0.0, 0.0],
        [22.0, 0.0, 0.0],
        [56.0, 0.0, 0.0],
        [60.0, 0.0, 0.0],
        [62.0, 0.0, 0.0],
        [40.0, 0.0, 0.0],
        [34.0, 0.0, 0.0],
    ],
    "BattleReady": [
        [14.0, 0.0, 0.0],
        [18.0, 0.0, 0.0],
        [28.0, 12.0, 0.0],
        [62.0, 36.0, 0.0],
        [66.0, 40.0, 0.0],
        [68.0, 42.0, 0.0],
        [46.0, 22.0, 0.0],
        [40.0, 18.0, 0.0],
    ],
    "Empowered": [
        [20.0, 10.0, 0.0],
        [24.0, 14.0, 0.0],
        [34.0, 18.0, 6.0],
        [68.0, 44.0, 22.0],
        [72.0, 48.0, 24.0],
        [74.0, 50.0, 24.0],
        [52.0, 28.0, 14.0],
        [46.0, 24.0, 10.0],
    ],
    "Overcharged": [
        [26.0, 14.0, 0.0],
        [30.0, 18.0, 0.0],
        [40.0, 22.0, 8.0],
        [74.0, 50.0, 28.0],
        [78.0, 54.0, 30.0],
        [80.0, 56.0, 30.0],
        [58.0, 34.0, 18.0],
        [52.0, 30.0, 14.0],
    ],
    "Cataclysmic": [
        [34.0, 20.0, 8.0],
        [40.0, 24.0, 10.0],
        [48.0, 28.0, 12.0],
        [82.0, 60.0, 36.0],
        [86.0, 64.0, 38.0],
        [88.0, 66.0, 38.0],
        [66.0, 40.0, 24.0],
        [58.0, 34.0, 18.0],
    ],
}


FACTION_COLORS = {
    "combat": "E8EEF7",
    "outlaws": "FDECEA",
    "wildfolk": "E8F5E9",
    "eraden": "FFF3E0",
    "eye": "EDE7F6",
    "rakta": "FCE4EC",
    "special": "E1F5FE",
    "fallback": "F5F5F5",
}


THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E78")
BODY_FONT = Font(name="Calibri", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def normalize_triplet(v1: float, v2: float, v3: float) -> tuple[float, float, float]:
    a = max(0.0, min(100.0, v1))
    b = max(0.0, min(100.0, v2))
    c = max(0.0, min(100.0, v3))
    total = a + b + c
    if total > 100.0 and total > 0.0001:
        scale = 100.0 / total
        a *= scale
        b *= scale
        c *= scale
    return round(a, 1), round(b, 1), round(c, 1)


def style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_body(ws, start_row: int = 2) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = LEFT


def color_row(ws, row: int, color_hex: str) -> None:
    fill = PatternFill("solid", fgColor=color_hex)
    for cell in ws[row]:
        cell.fill = fill


def set_widths(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width


def add_menu_workbook() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu Layout"

    ws.append(["Order", "Section", "Option", "Control", "Values", "Default", "Preset Writes", "Notes"])
    style_header(ws, 1)

    base_rows = [
        ["0", "Root (No Collapsible)", "Enable Mod", "Toggle", "On/Off", "On", "No", "Master switch"],
        ["5", "Factioned Imbuement", "Faction Profile Preset", "Dropdown", "Default | Core Factions | Most Factions | All Factions | Random", "Default", "N/A", "Batch-writes faction Enabled toggles"],
        ["7", "Factioned Imbuement", "Enemy Type Profile Preset", "Dropdown", "Mage | Mage Bow | Mage Melee | Mage Bow Melee | Random", "Mage", "N/A", "Batch-writes enemy archetype eligibility toggles"],
        ["10", "Factioned Imbuement", "Imbue Experience Preset", "Dropdown", "Default | Two-Slot | Tri-Slot | Tri-Slot+ | Random", "Default", "N/A", "Batch-writes faction Imbue 1/2/3"],
        ["20", "Factioned Imbuement", "Chance Experience Preset", "Dropdown", "Default | Increased | High | Very High | Maximum", "Default", "N/A", "Batch-writes faction Chance 1/2/3 (normalized)"],
        ["30", "Factioned Imbuement", "Strength Experience Preset", "Dropdown", "Default | Increased | High | Very High | Maximum", "Default", "N/A", "Batch-writes faction Strength 1/2/3"],
        ["40", "Enemy Type Eligibility", "Mage Eligible", "Toggle", "On/Off", "Preset-driven", "Enemy Type Profile", "Controls pure mage archetype eligibility"],
        ["41", "Enemy Type Eligibility", "Mage Bow Eligible", "Toggle", "On/Off", "Preset-driven", "Enemy Type Profile", "Controls mage-bow hybrid eligibility"],
        ["42", "Enemy Type Eligibility", "Mage Melee Eligible", "Toggle", "On/Off", "Preset-driven", "Enemy Type Profile", "Controls mage-melee hybrid eligibility"],
        ["43", "Enemy Type Eligibility", "Bow Eligible", "Toggle", "On/Off", "Preset-driven", "Enemy Type Profile", "Controls non-caster bow archetype eligibility"],
        ["44", "Enemy Type Eligibility", "Melee Eligible", "Toggle", "On/Off", "Preset-driven", "Enemy Type Profile", "Controls non-caster melee archetype eligibility"],
        ["45", "Enemy Type Eligibility", "Uncertain Enemy Type Fallback", "Dropdown", "Treat As Melee | Skip Enemy", "Treat As Melee", "No", "Controls behavior when enemy-type signals are uncertain"],
    ]
    for row in base_rows:
        ws.append(row)

    order_start = 100
    for i, (_, faction_name, short_name) in enumerate(FACTIONS):
        order = order_start + i * 10
        ws.append([str(order), faction_name, f"{short_name} Enabled", "Toggle", "On/Off", "Preset-driven", "Profile", "Source of truth for faction eligibility"])
        for slot in range(1, 4):
            ws.append([str(order + slot * 3 - 2), faction_name, f"{short_name} Imbue {slot}", "Dropdown", "None (No Imbue) + spell ids", "Preset-driven", "Imbue", "Source of truth for slot spell"])
            ws.append([str(order + slot * 3 - 1), faction_name, f"{short_name} Chance {slot}", "Slider", "0..100 in 5% steps", "Preset-driven", "Chance", "Source of truth for slot chance"])
            ws.append([str(order + slot * 3), faction_name, f"{short_name} Strength {slot}", "Slider", "0..100 in 5% steps", "Preset-driven", "Strength", "Source of truth for slot strength"])

    diag_rows = [
        ["999", "Advanced", "Basic Logs", "Toggle", "On/Off", "On", "No", "General informational logging"],
        ["1000", "Advanced", "Diagnostics Logs", "Toggle", "On/Off", "Off", "No", "Deeper troubleshooting logs"],
        ["1001", "Advanced", "Verbose Logs", "Toggle", "On/Off", "Off", "No", "High-volume per-creature logs"],
        ["1002", "Advanced", "Session Diagnostics", "Toggle", "On/Off", "Off", "No", "Structured session summary logs"],
        ["1003", "Advanced", "Imbue Update Interval", "Dropdown", "0.05s..1.00s", "0.25s", "No", "Performance/response tradeoff"],
        ["1004", "Advanced", "Enemy Rescan Interval", "Dropdown", "0.50s..5.00s", "2.00s", "No", "Tracking refresh interval"],
        ["1005", "Advanced", "Force Reapply", "Button", "False/True", "False", "No", "One-shot reapply action"],
        ["1010", "Advanced - Dumps", "Dump Factions", "Button", "False/True", "False", "No", "One-shot dump action"],
        ["1011", "Advanced - Dumps", "Dump Wave-Faction Map", "Button", "False/True", "False", "No", "One-shot dump action"],
        ["1012", "Advanced - Dumps", "Dump State", "Button", "False/True", "False", "No", "One-shot dump action"],
        ["1013", "Advanced - Dumps", "Dump Enemy Type Detection", "Button", "False/True", "False", "No", "One-shot dump action"],
    ]
    for row in diag_rows:
        ws.append(row)

    style_body(ws, 2)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    set_widths(ws, [9, 34, 30, 13, 52, 16, 12, 54])

    for row_idx in range(2, ws.max_row + 1):
        faction_name = ws.cell(row=row_idx, column=2).value
        for key, display, _ in FACTIONS:
            if faction_name == display:
                color_row(ws, row_idx, FACTION_COLORS[key])
                break

    impact = wb.create_sheet("Preset Impact")
    impact.append(["Section", "Faction", "Collapsible Option", "Source Of Truth", "Written By Preset Family", "Profile Can Disable", "Notes"])
    style_header(impact, 1)

    impact.append(["Enemy Type Eligibility", "N/A", "Mage Eligible", "Yes", "Enemy Type Profile", "N/A", "Runtime pure mage detection gate"])
    impact.append(["Enemy Type Eligibility", "N/A", "Mage Bow Eligible", "Yes", "Enemy Type Profile", "N/A", "Runtime mage-bow detection gate"])
    impact.append(["Enemy Type Eligibility", "N/A", "Mage Melee Eligible", "Yes", "Enemy Type Profile", "N/A", "Runtime mage-melee detection gate"])
    impact.append(["Enemy Type Eligibility", "N/A", "Bow Eligible", "Yes", "Enemy Type Profile", "N/A", "Runtime non-caster bow detection gate"])
    impact.append(["Enemy Type Eligibility", "N/A", "Melee Eligible", "Yes", "Enemy Type Profile", "N/A", "Runtime non-caster melee detection gate"])
    impact.append(["Enemy Type Eligibility", "N/A", "Uncertain Enemy Type Fallback", "Yes", "No", "N/A", "Fallback behavior for uncertain enemy-type detection"])

    for key, faction_name, short_name in FACTIONS:
        impact.append([faction_name, short_name, f"{short_name} Enabled", "Yes", "Faction Profile", "Yes", "Eligibility toggle controlled by Faction Profile Preset"])
        for slot in range(1, 4):
            impact.append([faction_name, short_name, f"{short_name} Imbue {slot}", "Yes", "Imbue", "Indirect", "None means no imbue slot"])
            impact.append([faction_name, short_name, f"{short_name} Chance {slot}", "Yes", "Chance", "Indirect", "Per-faction totals normalize to <=100%"])
            impact.append([faction_name, short_name, f"{short_name} Strength {slot}", "Yes", "Strength", "Indirect", "0..100 strength cap"])

    style_body(impact, 2)
    impact.freeze_panes = "A2"
    impact.auto_filter.ref = f"A1:G{impact.max_row}"
    set_widths(impact, [34, 12, 28, 16, 24, 16, 50])

    for row_idx in range(2, impact.max_row + 1):
        short_name = impact.cell(row=row_idx, column=2).value
        for key, _, short in FACTIONS:
            if short_name == short:
                color_row(impact, row_idx, FACTION_COLORS[key])
                break

    wb.save(MENU_PATH)


def add_matrix_workbook() -> None:
    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"

    overview.append(["Factioned Imbuement - Preset To Collapsible Mapping"])
    overview.append(["Source of truth: faction collapsible values. Presets only batch-write those values."])
    overview.append([""])
    overview.append(["Preset Family", "Writes Which Collapsible Options", "Rows Affected", "Notes"])
    overview.append(["Faction Profile", "Faction Enabled toggles", "8", "Controls which factions can roll imbues"])
    overview.append(["Enemy Type Profile", "Enemy archetype eligibility toggles", "5", "Controls which runtime enemy types are eligible"])
    overview.append(["Imbue", "Faction Imbue 1/2/3", "24", "Can write None to disable a slot"])
    overview.append(["Chance", "Faction Chance 1/2/3", "24", "Per-faction totals normalized to <=100%"])
    overview.append(["Strength", "Faction Strength 1/2/3", "24", "Clamped to 0..100"])
    overview.append([""])
    overview.append(["Total collapsible values managed by presets", "77", "5 eligibility toggles + 72 slot fields", ""])

    for cell in overview[1]:
        cell.font = TITLE_FONT
    style_header(overview, 4)
    style_body(overview, 5)
    set_widths(overview, [38, 48, 28, 54])

    profile_ws = wb.create_sheet("Faction Profile")
    profile_ws.append(["Faction"] + [label for _, label in FACTION_PROFILE_PRESETS] + ["Notes"])
    style_header(profile_ws, 1)
    for idx, (key, faction_name, _) in enumerate(FACTIONS):
        row = [faction_name]
        for preset_id, _ in FACTION_PROFILE_PRESETS:
            if preset_id == "RandomizedEligibility":
                row.append("Random each apply")
            else:
                row.append("Enabled" if FACTION_PROFILE_VALUES[preset_id][idx] else "Disabled")
        row.append("Eligibility only. Imbue/chance/strength rules still come from their own presets.")
        profile_ws.append(row)
    style_body(profile_ws, 2)
    profile_ws.freeze_panes = "A2"
    profile_ws.auto_filter.ref = f"A1:G{profile_ws.max_row}"
    set_widths(profile_ws, [34, 17, 17, 17, 19, 24, 60])
    for row_idx, (key, _, _) in enumerate(FACTIONS, start=2):
        color_row(profile_ws, row_idx, FACTION_COLORS[key])

    enemy_type_profile_ws = wb.create_sheet("Enemy Type Profile")
    enemy_type_profile_ws.append(["Enemy Type Toggle"] + [label for _, label in ENEMY_TYPE_PROFILE_PRESETS] + ["Notes"])
    style_header(enemy_type_profile_ws, 1)

    rows = [
        ("Mage Eligible", 0, "Enabled in all non-random presets."),
        ("Mage Bow Eligible", 1, "Enabled in all non-random presets."),
        ("Mage Melee Eligible", 2, "Enabled in all non-random presets."),
        ("Bow Eligible", 3, "Enabled by Mage Bow / Mage Bow Melee presets."),
        ("Melee Eligible", 4, "Enabled by Mage Melee / Mage Bow Melee presets."),
    ]

    for label, value_index, note in rows:
        row = [label]
        for preset_id, _ in ENEMY_TYPE_PROFILE_PRESETS:
            if preset_id == "RandomizedEligibility":
                row.append("Random each apply")
            else:
                row.append("On" if ENEMY_TYPE_PROFILE_VALUES[preset_id][value_index] else "Off")
        row.append(note)
        enemy_type_profile_ws.append(row)

    style_body(enemy_type_profile_ws, 2)
    enemy_type_profile_ws.freeze_panes = "A2"
    enemy_type_profile_ws.auto_filter.ref = f"A1:G{enemy_type_profile_ws.max_row}"
    set_widths(enemy_type_profile_ws, [34, 20, 22, 20, 20, 24, 66])

    imbue_ws = wb.create_sheet("Imbue Writes")
    imbue_ws.append(["Faction", "Collapsible Option", "Slot"] + [label for _, label in IMBUE_PRESETS] + ["Notes"])
    style_header(imbue_ws, 1)
    for i, (key, faction_name, short_name) in enumerate(FACTIONS):
        for slot in range(1, 4):
            row = [faction_name, f"{short_name} Imbue {slot}", f"Slot {slot}"]
            for preset_id, _ in IMBUE_PRESETS:
                if preset_id == "Randomized":
                    row.append("Random Fire/Lightning/Gravity")
                else:
                    row.append(IMBUE_VALUES[preset_id][i][slot - 1])
            row.append("Exact value written by Imbue preset")
            imbue_ws.append(row)
    style_body(imbue_ws, 2)
    imbue_ws.freeze_panes = "A2"
    imbue_ws.auto_filter.ref = f"A1:I{imbue_ws.max_row}"
    set_widths(imbue_ws, [34, 24, 10, 16, 18, 18, 18, 20, 44])
    for row_idx in range(2, imbue_ws.max_row + 1):
        faction_name = imbue_ws.cell(row=row_idx, column=1).value
        for key, display, _ in FACTIONS:
            if faction_name == display:
                color_row(imbue_ws, row_idx, FACTION_COLORS[key])
                break

    chance_ws = wb.create_sheet("Chance Writes")
    chance_ws.append(["Faction", "Collapsible Option", "Slot"] + [label for _, label in CHANCE_PRESETS] + ["Notes"])
    style_header(chance_ws, 1)
    for i, (key, faction_name, short_name) in enumerate(FACTIONS):
        for slot in range(1, 4):
            row = [faction_name, f"{short_name} Chance {slot}", f"Slot {slot}"]
            for preset_id, _ in CHANCE_PRESETS:
                v = CHANCE_VALUES[preset_id][i]
                n1, n2, n3 = normalize_triplet(v[0], v[1], v[2])
                row.append(f"{[n1, n2, n3][slot - 1]:.1f}%")
            row.append("Written value after per-faction normalization")
            chance_ws.append(row)
    style_body(chance_ws, 2)
    chance_ws.freeze_panes = "A2"
    chance_ws.auto_filter.ref = f"A1:I{chance_ws.max_row}"
    set_widths(chance_ws, [34, 24, 10, 16, 18, 18, 20, 16, 48])
    for row_idx in range(2, chance_ws.max_row + 1):
        faction_name = chance_ws.cell(row=row_idx, column=1).value
        for key, display, _ in FACTIONS:
            if faction_name == display:
                color_row(chance_ws, row_idx, FACTION_COLORS[key])
                break

    strength_ws = wb.create_sheet("Strength Writes")
    strength_ws.append(["Faction", "Collapsible Option", "Slot"] + [label for _, label in STRENGTH_PRESETS] + ["Notes"])
    style_header(strength_ws, 1)
    for i, (key, faction_name, short_name) in enumerate(FACTIONS):
        for slot in range(1, 4):
            row = [faction_name, f"{short_name} Strength {slot}", f"Slot {slot}"]
            for preset_id, _ in STRENGTH_PRESETS:
                row.append(f"{STRENGTH_VALUES[preset_id][i][slot - 1]:.1f}%")
            row.append("Written value after 0..100 clamp")
            strength_ws.append(row)
    style_body(strength_ws, 2)
    strength_ws.freeze_panes = "A2"
    strength_ws.auto_filter.ref = f"A1:I{strength_ws.max_row}"
    set_widths(strength_ws, [34, 24, 10, 16, 18, 18, 20, 16, 48])
    for row_idx in range(2, strength_ws.max_row + 1):
        faction_name = strength_ws.cell(row=row_idx, column=1).value
        for key, display, _ in FACTIONS:
            if faction_name == display:
                color_row(strength_ws, row_idx, FACTION_COLORS[key])
                break

    wb.save(MATRIX_PATH)


def main() -> None:
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    add_menu_workbook()
    add_matrix_workbook()
    print(f"Generated {MENU_PATH}")
    print(f"Generated {MATRIX_PATH}")


if __name__ == "__main__":
    main()
